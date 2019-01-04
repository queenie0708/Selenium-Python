from selenium import webdriver
from openpyxl import Workbook
from openpyxl import load_workbook
# def createExcel():
#     book = Workbook()
#     sheetNames = ['CTS','GTS','VTS','GSI']
#     for sheetName in sheetNames:
#       sheet = book.create_sheet(0)
#       sheet.title = sheetName
#       sheet["A1"] = 'No.' # 其中的'0-行, 0-列'指定表中的单元
#       sheet['B1'] ='Module'
#       sheet['C1'] ='Test'
#       sheet['D1'] ='Result'
#       sheet['E1'] ='Details'
#       sheet['F1'] ='ErrorID:'
#    book.save(r'D:\test\test1.xlsx')  # 在字符串前加r，声明为raw字符串，这样就不会处理其中的转义了。否则，可能会报错
book = load_workbook('d:\\ANT\\ANT_V0.270_CTS&GTSfailedcases.xlsx')# need to use \\ instead of \ 
browser = webdriver.Chrome()
browser.get('file://10.75.10.81/Share/ANT/CTS/V0.270/GTS-toNJ/results/2018.12.27_23.13.46/test_result_failures.html')
testNames = browser.find_elements_by_class_name('testname')
details = browser.find_elements_by_class_name('details')
i = 2
j = 2
sheet = book.get_sheet_by_name("GTS")
for test in testNames:
    sheet['C'+str(i)] = test.text
    i = i + 1
#   print(test.text)
# print('==========================================below is details===========================')
for detail in details:
    sheet['E'+str(j)] = detail.text
    j = j + 1
#   print(detail.text)
browser.quit()
book.save('d:\\ANT\\ANT_V0.270_CTS&GTSfailedcases.xlsx')
