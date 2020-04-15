from openpyxl import Workbook
from openpyxl import load_workbook
import xml.dom.minidom
import os
book = Workbook()
sheet = book.create_sheet(0)
sheet["A1"] = 'Google Submission' 
sheet['B1'] ='Progress'
sheet['C1'] ='Pass'
sheet['D1'] ='Failed'
sheet['E1'] ='Models'
sheet['F1'] ='Comment'

sheet["A10"] ='Google Submission' 
sheet['B10'] ='Test Suit'
sheet['C10'] ='Build'
sheet['D10'] ='Commend'
sheet['E10'] ='Retry'

Commend = {
               'CTS' : 'run cts  --shard-count x',
               'GSI' : 'run cts-on-gsi –shard-count x',
               'VTS' : 'run vts',
               'STS' : 'run sts-engbuild',
               'GTS' : 'run gts-suite',
               'ETS' : 'run ets'
               }
Retry = {
               'CTS' : 'run retry –retey x  –shard-count x',
               'GSI' : 'run cts-on-gsi-retry  –retry x',
               'VTS' : 'run vts –retry x',
               'STS' : 'run sts-engbuild –retry x',
               'GTS' : 'run retry –retry x',
               'ETS' : 'run ets'
        }

def getAttr(path,ele,attr):
   dom = xml.dom.minidom.parse(path + '/test_result.xml')
   root = dom.documentElement
   Ele = root.getElementsByTagName(ele)
   Attr = Ele[0].getAttribute(attr)
   return Attr
rootPath = '//10.75.10.81/Share/FRT/CTS_report/V1.520/00WW' #//10.75.10.81/Share/CO2P/Google_Submission/V3.110/SKU2
dirs = os.listdir(rootPath)
path = rootPath + '/' + dirs[0] + '/results'
subDir = os.listdir(path)
subPath = path + '/' + subDir[0]

for i in range(len(dirs)):
    testPath = rootPath + '/' + dirs[i] + '/results'
    resultPath = testPath + '/' + os.listdir(testPath)[0]
    print('now we are in ' + str(testPath))
    dom = xml.dom.minidom.parse(resultPath + '/test_result.xml')
    root = dom.documentElement
    suite = root.getAttribute('suite_version')
    print('suite =' + suite)
    sheet['B'+str(i+11)] = suite
    sheet['A'+str(i+2)] = dirs[i]
    sheet['A'+str(i+11)] = dirs[i]
    sheet['B'+str(i+2)] = '100%'
  
    build = getAttr(resultPath,'Build','build_type')
    print('build =' + build)
    sheet['C'+str(i+11)] = build
    passCount = getAttr(resultPath,'Summary','pass')
    print('pass =' + passCount)
    sheet['C'+str(i+2)] = passCount
    failedCount = getAttr(resultPath,'Summary','failed')
    print('failed =' + failedCount)
    sheet['D'+str(i+2)] = failedCount
    modelDone = getAttr(resultPath,'Summary','modules_done')
    print('modules done =' + modelDone)
    modelTotal = getAttr(resultPath,'Summary','modules_total')
    print('modules Total =' + modelTotal)
    sheet['E'+str(i+2)] = modelDone + '/' + modelTotal
    if failedCount == '0':
        comment = "pass"
    else:
        comment = failedCount + 'items'
    sheet['F'+str(i+2)] = comment
    if dirs[i] in Commend:
        commend = Commend[dirs[i]]
        print('commend = '+ commend)
        sheet['D'+str(i+11)] = commend
    if dirs[i] in Retry:
        retry = Retry[dirs[i]]
        print('retry = '+ retry)
        sheet['E'+str(i+11)] = retry
book.save(r'D:\xtsFRTpReport.xlsx')
