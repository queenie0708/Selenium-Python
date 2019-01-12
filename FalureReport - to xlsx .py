from selenium import webdriver
from openpyxl import Workbook
from openpyxl import load_workbook

# book = Workbook()
# sheetNames = ['CO2','ES2','E2M','EVW']
# for sheetName in sheetNames:
#   sheet = book.create_sheet(0)
#   sheet.title = sheetName
#   sheet["A1"] = 'No.' # 其中的'0-行, 0-列'指定表中的单元
#   sheet['B1'] ='Module'
#   sheet['C1'] ='Test'
#   sheet['D1'] ='Result'
#   sheet['E1'] ='Details'
# book.save('d:\\test\\result\\GTSfailedcases.xlsx')  # 在字符串前加r，声明为raw字符串，这样就不会处理其中的转义了。否则，可能会报错
book = load_workbook('d:\\test\\result\\GTSfailedcases.xlsx')# need to use \\ instead of \
browser = webdriver.Chrome()
browser.get('file:///D:/E2/GTS/results/2019.01.02_14.06.11/test_result_failures.html')
testNames = browser.find_elements_by_class_name('testname')
details = browser.find_elements_by_class_name('details')
i = 20
j = 20
sheet = book.get_sheet_by_name("ES2P")
for test in testNames:
    sheet['C'+str(i)] = test.text
    sheet['A'+str(i)] = 'V3.110_SKU2'
    i = i + 1
#   print(test.text)
# print('==========================================below is details===========================')
for detail in details:
    sheet['E'+str(j)] = detail.text
    j = j + 1
#   print(detail.text)
browser.quit()
book.save('d:\\test\\result\\GTSfailedcases.xlsx')
